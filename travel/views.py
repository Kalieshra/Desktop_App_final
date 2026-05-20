from datetime import datetime, timedelta
from django.utils.timezone import localdate

import openpyxl
from openpyxl.styles import Font

from django.conf import settings
from django.contrib import messages
from django.db.models import Q, Count, OuterRef, Subquery, IntegerField
from django.http import FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views import View
from django.views.decorators.cache import never_cache
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView

from .forms import TravelForm, ExcelImportForm, TravelTypeForm
from .models import Travel, TravelType


@never_cache
def dashboard(request):
    q = request.GET.get('q', '').strip()
    context = {'q': q}

    if q:
        from licenses.models import Channel, PersonalLicense, Inspection
        from supplies.models import SupplyChannel, BuyOrder, Offer

        person_count_sq = (
            Travel.objects.filter(name=OuterRef('name'))
            .values('name')
            .annotate(c=Count('id'))
            .values('c')
        )
        context['travel_results'] = Travel.objects.filter(
            Q(name__icontains=q) |
            Q(course_name__icontains=q) |
            Q(course_location__icontains=q) |
            Q(event_code__icontains=q)
        ).annotate(
            person_travels_count=Subquery(person_count_sq, output_field=IntegerField())
        )[:10]

        context['license_channels'] = Channel.objects.filter(
            Q(name__icontains=q)
        )[:10]
        context['personal_licenses'] = PersonalLicense.objects.filter(
            Q(applicant_name__icontains=q)
        )[:10]
        context['inspections'] = Inspection.objects.filter(
            Q(inspector_name__icontains=q)
        )[:10]

        context['supply_channels'] = SupplyChannel.objects.filter(
            Q(name__icontains=q)
        )[:10]
        context['buy_orders'] = BuyOrder.objects.filter(
            Q(requester_name__icontains=q)
        )[:10]
        context['offers'] = Offer.objects.filter(
            Q(company_name__icontains=q)
        )[:10]
    else:
        from licenses.models import Channel, PlaceLicense, PersonalLicense, Inspection
        from supplies.models import SupplyChannel, BuyOrder, Offer

        # Travel stats
        travel_total = Travel.objects.count()
        travel_type_stats = (
            Travel.objects.filter(travel_type__isnull=False)
            .values('travel_type__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        department_stats = (
            Travel.objects.exclude(department='')
            .values('department')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        person_count_sq = (
            Travel.objects.filter(name=OuterRef('name'))
            .values('name')
            .annotate(c=Count('id'))
            .values('c')
        )
        recent_travel = Travel.objects.annotate(
            person_travels_count=Subquery(person_count_sq, output_field=IntegerField())
        )[:10]

        today = localdate()
        alert_cutoff = today + timedelta(days=9)
        deadline_alert_travels = Travel.objects.filter(
            deadline__gte=today,
            deadline__lte=alert_cutoff,
        ).order_by('deadline')
        deadline_alert_count = deadline_alert_travels.count()

        # Licenses stats
        license_channels_total = Channel.objects.count()
        place_licenses_total = PlaceLicense.objects.count()
        personal_licenses_total = PersonalLicense.objects.count()
        inspections_total = Inspection.objects.count()
        recent_license_channels = Channel.objects.all()[:10]

        # Supplies stats
        supply_channels_total = SupplyChannel.objects.count()
        buy_orders_total = BuyOrder.objects.count()
        offers_total = Offer.objects.count()
        recent_supply_channels = SupplyChannel.objects.all()[:10]

        context.update({
            'travel_total': travel_total,
            'travel_type_stats': travel_type_stats,
            'department_stats': department_stats,
            'recent_travel': recent_travel,
            'deadline_alert_count': deadline_alert_count,
            'deadline_alert_travels': deadline_alert_travels,
            # Licenses
            'license_channels_total': license_channels_total,
            'place_licenses_total': place_licenses_total,
            'personal_licenses_total': personal_licenses_total,
            'inspections_total': inspections_total,
            'recent_license_channels': recent_license_channels,
            # Supplies
            'supply_channels_total': supply_channels_total,
            'buy_orders_total': buy_orders_total,
            'offers_total': offers_total,
            'recent_supply_channels': recent_supply_channels,
        })

    return render(request, 'dashboard.html', context)


class TravelListView(ListView):
    model = Travel
    template_name = 'travel/travel_list.html'
    context_object_name = 'travels'
    paginate_by = 25

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q', '').strip()
        travel_type = self.request.GET.get('travel_type', '').strip()
        department = self.request.GET.get('department', '').strip()

        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(course_name__icontains=q) |
                Q(course_location__icontains=q) |
                Q(event_code__icontains=q)
            )
        if travel_type:
            qs = qs.filter(travel_type_id=travel_type)
        if department:
            qs = qs.filter(department=department)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get('q', '')
        context['selected_travel_type'] = self.request.GET.get('travel_type', '')
        context['selected_department'] = self.request.GET.get('department', '')
        context['travel_type_choices'] = TravelType.objects.filter(is_active=True)
        context['department_choices'] = Travel.DEPARTMENT_CHOICES
        return context


class TravelByPersonListView(ListView):
    model = Travel
    template_name = 'travel/travel_by_person.html'
    context_object_name = 'travels'
    paginate_by = 25

    def get_queryset(self):
        self.person_name = self.request.GET.get('name', '').strip()
        if not self.person_name:
            return Travel.objects.none()
        return Travel.objects.filter(name=self.person_name).order_by('-date_from', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['person_name'] = self.person_name
        context['travels_count'] = self.get_queryset().count()
        return context


class TravelDetailView(DetailView):
    model = Travel
    template_name = 'travel/travel_detail.html'
    context_object_name = 'travel'


class TravelCreateView(CreateView):
    model = Travel
    form_class = TravelForm
    template_name = 'travel/travel_form.html'
    success_url = reverse_lazy('travel:list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة السجل بنجاح.')
        return super().form_valid(form)


class TravelUpdateView(UpdateView):
    model = Travel
    form_class = TravelForm
    template_name = 'travel/travel_form.html'
    success_url = reverse_lazy('travel:list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تعديل السجل بنجاح.')
        return super().form_valid(form)


class TravelDeleteView(DeleteView):
    model = Travel
    template_name = 'travel/travel_confirm_delete.html'
    success_url = reverse_lazy('travel:list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف السجل بنجاح.')
        return super().form_valid(form)


def _parse_date(value):
    """Parse date from various Excel formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    # Excel serial number
    if isinstance(value, (int, float)):
        try:
            from datetime import timedelta, date
            base = date(1899, 12, 30)
            return base + timedelta(days=int(value))
        except (ValueError, OverflowError):
            pass
    return None


def import_excel(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded = request.FILES['file']
            wb = openpyxl.load_workbook(uploaded, data_only=True)
            ws = wb.worksheets[0]

            count = 0
            for row in ws.iter_rows(min_row=9, values_only=True):
                # Skip empty rows
                if not row[1]:
                    continue

                Travel.objects.create(
                    name=str(row[1] or '').strip(),
                    title=str(row[2] or '').strip(),
                    department=str(row[3] or '').strip(),
                    position=str(row[4] or '').strip(),
                    course_type=str(row[5] or '').strip(),
                    course_name=str(row[6] or '').strip(),
                    course_location=str(row[7] or '').strip(),
                    date_from=_parse_date(row[8]),
                    date_to=_parse_date(row[9]),
                    deadline=_parse_date(row[10]),
                    followup=str(row[11] or '').strip(),
                    event_code=str(row[13] or '').strip(),
                    total_travels=str(row[14] or '').strip(),
                )
                count += 1

            messages.success(request, f'تم استيراد {count} سجل بنجاح.')
            return redirect('travel:list')
    else:
        form = ExcelImportForm()
    return render(request, 'travel/travel_import.html', {'form': form})


class TravelToggleAcceptView(View):
    def post(self, request, pk):
        obj = get_object_or_404(Travel, pk=pk)
        obj.is_accepted = not obj.is_accepted
        obj.save()
        messages.success(request, 'تم تحديث حالة السفرية بنجاح.')
        return redirect('travel:list')


# ──────────────────────────────────────────────────────────
# TravelType CRUD Views
# ──────────────────────────────────────────────────────────

class TravelTypeListView(ListView):
    model = TravelType
    template_name = 'travel/travel_type_list.html'
    context_object_name = 'travel_types'
    paginate_by = 25

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get('q', '')
        return context


class TravelTypeDetailView(DetailView):
    model = TravelType
    template_name = 'travel/travel_type_detail.html'
    context_object_name = 'travel_type'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Show all travels using this type
        context['travels'] = self.object.travels.all()[:10]
        context['travels_count'] = self.object.travels.count()
        return context


class TravelTypeCreateView(CreateView):
    model = TravelType
    form_class = TravelTypeForm
    template_name = 'travel/travel_type_form.html'
    success_url = reverse_lazy('travel:travel_type_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة نوع السفرية بنجاح.')
        return super().form_valid(form)


class TravelTypeUpdateView(UpdateView):
    model = TravelType
    form_class = TravelTypeForm
    template_name = 'travel/travel_type_form.html'
    success_url = reverse_lazy('travel:travel_type_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تعديل نوع السفرية بنجاح.')
        return super().form_valid(form)


class TravelTypeDeleteView(DeleteView):
    model = TravelType
    template_name = 'travel/travel_type_confirm_delete.html'
    success_url = reverse_lazy('travel:travel_type_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف نوع السفرية بنجاح.')
        return super().form_valid(form)


def backup_all(request):
    from licenses.models import Channel, PlaceLicense, PersonalLicense, Inspection
    from supplies.models import SupplyChannel, BuyOrder, ServiceOrder, Offer, ChannelPost

    today = datetime.now()
    date_folder = today.strftime("%Y-%m-%d")
    backup_dir = settings.BASE_DIR / 'backup' / date_folder
    backup_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
        for row in rows:
            ws.append(row)

    add_sheet('TravelType',
        ['ID', 'الاسم', 'الوصف', 'نشط', 'تاريخ الإنشاء'],
        [[o.id, o.name, o.description, o.is_active, str(o.created_at)]
         for o in TravelType.objects.all()])

    add_sheet('Travel',
        ['ID', 'الاسم', 'اللقب', 'الجهة', 'الوظيفة', 'نوع السفرية', 'نوع الدورة',
         'اسم الدورة', 'مكان الدورة', 'من', 'إلى', 'Deadline', 'المتابعة',
         'Event Code', 'إجمالي السفريات', 'مقبول', 'APS', 'لجنة النشر',
         'عرض مرئي', 'صادر', 'تاريخ الإنشاء'],
        [[o.id, o.name, o.title, o.department, o.position,
          str(o.travel_type) if o.travel_type else '',
          o.course_type, o.course_name, o.course_location,
          str(o.date_from), str(o.date_to), str(o.deadline), o.followup,
          o.event_code, o.total_travels, o.is_accepted, o.aps,
          o.publish_committee, o.visual_presentation, o.sader, str(o.created_at)]
         for o in Travel.objects.select_related('travel_type').all()])

    add_sheet('قنوات التراخيص',
        ['ID', 'الاسم', 'التاريخ', 'أنشئ بواسطة', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, o.name, str(o.date), o.created_by, o.is_accepted, str(o.created_at)]
         for o in Channel.objects.all()])

    add_sheet('التراخيص المكانية',
        ['ID', 'القناة', 'مدة الترخيص', 'من', 'إلى', 'تم السداد', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, str(o.channel), o.license_duration, str(o.date_from), str(o.date_to),
          o.is_paid, o.is_accepted, str(o.created_at)]
         for o in PlaceLicense.objects.select_related('channel').all()])

    add_sheet('التراخيص الشخصية',
        ['ID', 'القناة', 'اسم المتقدم', 'الفئة', 'الوظيفة', 'الهاتف',
         'نوع الترخيص', 'من', 'إلى', 'مدة الترخيص', 'تم السداد', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, str(o.channel), o.applicant_name, o.category, o.job, o.phone,
          o.license_type, str(o.date_from), str(o.date_to), o.license_duration,
          o.is_paid, o.is_accepted, str(o.created_at)]
         for o in PersonalLicense.objects.select_related('channel').all()])

    add_sheet('التفتيش',
        ['ID', 'القناة', 'اسم المفتش', 'الوظيفة', 'جهة التفتيش',
         'تاريخ الزيارة', 'النتيجة', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, str(o.channel), o.inspector_name, o.job, o.authority,
          str(o.visit_date), o.result, o.is_accepted, str(o.created_at)]
         for o in Inspection.objects.select_related('channel').all()])

    add_sheet('قنوات التوريدات',
        ['ID', 'الاسم', 'التاريخ', 'أنشئ بواسطة', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, o.name, str(o.date), o.created_by, o.is_accepted, str(o.created_at)]
         for o in SupplyChannel.objects.all()])

    add_sheet('طلبات الشراء',
        ['ID', 'القناة', 'اسم مقدم الطلب', 'الوظيفة', 'المكان',
         'البيان', 'القيمة التقديرية', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, str(o.supply_channel), o.requester_name, o.job, o.location,
          o.description, str(o.estimated_value), o.is_accepted, str(o.created_at)]
         for o in BuyOrder.objects.select_related('supply_channel').all()])

    add_sheet('أوامر الخدمة',
        ['ID', 'طلب الشراء', 'رقم الأمر', 'التاريخ', 'الموضوع',
         'إلى', 'البنود', 'التوقيع', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, str(o.buy_order), o.service_order_no, str(o.date),
          o.object_description, o.to_field, o.items, o.signature,
          o.is_accepted, str(o.created_at)]
         for o in ServiceOrder.objects.select_related('buy_order').all()])

    add_sheet('العروض',
        ['ID', 'طلب الشراء', 'اسم الشركة', 'السعر', 'الوصف',
         'مختار', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, str(o.buy_order), o.company_name, str(o.price),
          o.description, o.is_selected, o.is_accepted, str(o.created_at)]
         for o in Offer.objects.select_related('buy_order').all()])

    add_sheet('المنشورات',
        ['ID', 'القناة', 'العنوان', 'التاريخ', 'الوصف',
         'السعر', 'مقبول', 'تاريخ الإنشاء'],
        [[o.id, str(o.supply_channel), o.name, str(o.date),
          o.description, str(o.price), o.is_accepted, str(o.created_at)]
         for o in ChannelPost.objects.select_related('supply_channel').all()])

    filename = f'backup_{today.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    filepath = backup_dir / filename
    wb.save(filepath)

    messages.success(request, f'تم حفظ النسخة الاحتياطية: backup/{date_folder}/{filename}')
    return redirect('travel:dashboard')
